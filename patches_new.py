from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from datetime import datetime
import pandas as pd
import re
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils.cell import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink

# Script definitions
PREVIOUS_PATCH_WEDNESDAY = "2024-08-14"

# URLs for Microsoft Catalog patch gathering
catalog_urls = [
    ("https://www.catalog.update.microsoft.com/Search.aspx?q=Windows+Server+2016+for+x64+2024","Windows Server 2016 (2024)","version 1607"),
    ("https://www.catalog.update.microsoft.com/Search.aspx?q=Windows+Server+2019+for+x64+2024","Windows Server 2019 (2024)","version 1809"),
    ("https://www.catalog.update.microsoft.com/Search.aspx?q=Windows+Server+2016+for+x64","Windows Server 2016","version 1607"),
    ("https://www.catalog.update.microsoft.com/Search.aspx?q=Windows+Server+2019+for+x64","Windows Server 2019","version 1809"),
    ("https://www.catalog.update.microsoft.com/Search.aspx?q=Microsoft+Server+Operating+System-21H2+for+x64","Windows Server 2022","version 21H2"),
    ("https://www.catalog.update.microsoft.com/Search.aspx?q=Windows+10+Version+22H2+for+x64","Windows 10","version 22H2"),
    ("https://www.catalog.update.microsoft.com/Search.aspx?q=Windows+11+Version+23H2+for+x64","Windows 11","version 23H2"),
    ("https://www.catalog.update.microsoft.com/Search.aspx?q=SQL+Server+2016+Service+Pack+3","SQL Server 2016","SP3 (13.0.6419.1)"),
    ("https://www.catalog.update.microsoft.com/Search.aspx?q=SQL+Server+2019","SQL Server 2019","RTM"),
    ("https://www.catalog.update.microsoft.com/Search.aspx?q=Microsoft+Office+2016+32-bit","Office 2016","x32 bits")
] 

def catalog_scrape(urls, cutoff_date, system1_csv, patchstatus_csv, header_mappings=None, remove_columns=None, filter_terms=None, screenshot=True):
    cutoff_date = datetime.strptime(cutoff_date, '%Y-%m-%d')
    service = Service(executable_path="chromedriver.exe")
    driver = webdriver.Chrome(service=service)

    sys1AllRows = []

    try:
        for url, label, version in urls:
            driver.get(url)
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "ctl00_catalogBody_updateMatches")))
            driver.find_element(By.ID, "ctl00_catalogBody_updateMatches_ctl02_dateHeaderLink").click()

            if screenshot:
                driver.save_screenshot(f'{label}_{datetime.now().strftime("%Y_%m_%d")}.png')

            table = driver.find_element(By.ID, "ctl00_catalogBody_updateMatches")
            rows = table.find_elements(By.TAG_NAME, "tr")

            header = [th.text.strip() for th in rows[0].find_elements(By.TAG_NAME, 'th')]
            header[0] = '#'

            column_indices = {name: index for index, name in enumerate(header)}

            if header_mappings:
                for key, value in header_mappings.items():
                    if key in column_indices and value in column_indices:
                        header[column_indices[key]], header[column_indices[value]] = header[column_indices[value]], header[column_indices[key]]

            if remove_columns:
                header = [col for col in header if col not in remove_columns]

            system1_rows = []
            patchstatus_rows = []
            patchNumber = 1

            for row in rows[1:]:
                cells = row.find_elements(By.TAG_NAME, 'td')
                if len(cells) > 0:
                    title_text = cells[column_indices['Title']].text
                    date_str = cells[column_indices['Last Updated']].text

                    try:
                        row_date = datetime.strptime(date_str, '%m/%d/%Y')
                        if row_date >= cutoff_date:
                            kb_search = re.search(r"KB\d{6,}", title_text)
                            kb_number = kb_search.group() if kb_search else "No KB Number"

                            new_row = [patchNumber] + [cells[column_indices[col]].text for col in header[1:] if col in column_indices] + \
                                      ["Microsoft", label, version, f"{kb_number}", f"https://support.microsoft.com/kb/{kb_number}" if kb_search else "No Link"]
                            patchNumber += 1

                            if filter_terms and any(term.lower() in title_text.lower() for term in filter_terms):
                                patchstatus_rows.append(new_row)
                                patchNumber -= 1
                            else:
                                system1_rows.append(new_row)
                                sys1AllRows.append(new_row)
                    except ValueError:
                        print(f"Invalid date format found in row {title_text} {date_str}")

            header.extend(['Vendor Name', 'Vendor Product', 'Model/Version', 'Patch Name', 'Patch Link'])
            header[1] = 'Patch Description'
            header[2] = 'Release Date'
            header[3] = 'Update Type'

    finally:
        sys1df = pd.DataFrame(data=sys1AllRows, columns=header)
        driver.quit()
        return sys1df

def build_excel():
    catalogDf = catalog_scrape(
        urls=catalog_urls,
        cutoff_date=PREVIOUS_PATCH_WEDNESDAY,
        system1_csv=f"system1_{datetime.now().strftime('%Y_%m_%d')}.csv",
        patchstatus_csv=f"patch_status_{datetime.now().strftime('%Y_%m_%d')}.csv",
        header_mappings={'Classification': 'Last Updated'},
        remove_columns=['Products', 'Version', 'Size', 'Download'],
        filter_terms=[
            'Preview', 'Dynamic', 'Azure Stack HCI', 'Office 2019', 'Access', 'Project',
            'Outlook', 'PowerPoint', 'Visio', 'Publisher', '3.5, 4.8 and 4.8.1', '3.5, 4.7.2 and 4.8',
            '.NET Framework 3.5 and 4.8.1 for Windows 10 Version 22H2 for x64'
        ],
        screenshot=False
    )

    masterDf = catalogDf
    header_list = ["#", "Vendor Name", "Vendor Product", "Model/Version", "Patch Name", "Patch Description", "Patch Link", "Release Date", "Update Type", "Approved by BN", "Test Status", "Comment"]
    masterDf = masterDf.reindex(columns=header_list)

    # Save DataFrame to Excel using openpyxl engine
    with pd.ExcelWriter("System 1 Patches.xlsx", engine="openpyxl") as writer:
        masterDf.to_excel(writer, sheet_name="September-2024", index=False)

        workbook = writer.book
        worksheet = workbook["September-2024"]

        # Set specific column and row dimensions
        worksheet.column_dimensions["A"].width = 3
        worksheet.column_dimensions["B"].width = 7
        worksheet.column_dimensions["C"].width = 11
        worksheet.column_dimensions["D"].width = 12
        worksheet.column_dimensions["E"].width = 11
        worksheet.column_dimensions["F"].width = 45
        worksheet.column_dimensions["G"].width = 30
        worksheet.column_dimensions["H"].width = 10
        worksheet.column_dimensions["I"].width = 11
        worksheet.column_dimensions["J"].width = 7
        worksheet.column_dimensions["K"].width = 7
        worksheet.column_dimensions["L"].width = 35
        worksheet.row_dimensions[1].height = 46.5
        worksheet.row_dimensions[2].height = 46.5
        worksheet.row_dimensions[3].height = 34.5

        # Define the styles
        header_fill = PatternFill(fill_type="solid", start_color="DDEBF7", end_color="DDEBF7")
        header_font = Font(name='Calibri', color="000000", size=8, bold=True)
        center_alignment = Alignment(horizontal="center", vertical="center",wrap_text=True)

        calibri_font_8 = Font(name='Calibri', size=8)
        calibri_hyperlink = Font(name='Calibri', size=8, color="0000FF", underline="single")
        green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")

        # Define thin border for all cells
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"), 
            top=Side(style="thin"), bottom=Side(style="thin")
        )
        # Style header row
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = thin_border

        # Style all other rows and apply hyperlink/green fill for column E
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                # Set font for all cells
                cell.font = calibri_font_8

                # Apply borders to all cells
                cell.border = thin_border

                # Check if cell contains a URL and convert to hyperlink
                if isinstance(cell.value, str) and cell.value.startswith("http"):
                    cell.hyperlink = cell.value  # Convert to hyperlink
                    cell.font = calibri_hyperlink  # Apply hyperlink font

                # Apply green fill for column E (excluding header)
                if cell.column == 5:  # Column E (Excel is 1-indexed)
                    cell.fill = green_fill

                # Set alignment: 
                # - Center alignment for all cells horizontally and vertically
                # - For columns F and G: vertically centered but horizontally left-aligned
                if cell.column in [6, 7]:  # Columns F and G
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                else:
                    cell.alignment = center_alignment  # Center horizontally and vertically for all other columns


build_excel()