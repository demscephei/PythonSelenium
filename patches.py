from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from datetime import datetime, timedelta
import pandas as pd
import re
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils.cell import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink
import time

# Script definitions
PREVIOUS_PATCH_WEDNESDAY = "2024-08-14"

def av_scrape():
    service = Service(executable_path="chromedriver.exe")
    driver = webdriver.Chrome(service=service)
    headers = ["#", "Patch Description", "Release Date", "Update Type", "Vendor Name", "Vendor Product", "Model/Version"]
    av_rows = []
    try:
        for url,vendor, product, description, searchfor in av_urls:
            # Go to AV defintion site
            driver.get(url)
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH(searchfor)))
            )
          
            new_row = ["1"] + [description] + ["todays date"] + ["Virus Def."] + [vendor] + [product] + ["ver"]
            av_rows.append(new_row)
    finally:
        avdf = pd.DataFrame(data=av_rows, columns=headers)
        driver.quit()
        print(avdf)
        #return avdf

# URLs for Microsoft Catalog patch gathering
av_urls = []

# URLs for Microsoft Catalog patch gathering
catalog_urls = [
    ("https://www.catalog.update.microsoft.com/Search.aspx?q=Windows+Server+2016+for+x64+2024","Windows Server 2016 (2024)","version 1607"),
    ("https://www.catalog.update.microsoft.com/Search.aspx?q=Windows+Server+2019+for+x64+2024","Windows Server 2019 (2024)","version 1809"),
    ("https://www.catalog.update.microsoft.com/Search.aspx?q=Windows+Server+2016+for+x64","Windows Server 2016","version 1607"),
    ("https://www.catalog.update.microsoft.com/Search.aspx?q=Windows+Server+2019+for+x64","Windows Server 2019","version 1809"),
    ("https://www.catalog.update.microsoft.com/Search.aspx?q=Microsoft+Server+Operating+System-21H2+for+x64","Windows Server 2022","version 21H2"),
    ("https://www.catalog.update.microsoft.com/Search.aspx?q=Windows+10+Version+22H2+for+x64","Windows 10","version 22H2"),
    ("https://www.catalog.update.microsoft.com/Search.aspx?q=Windows+11+Version+23H2+for+x64","Windows 11","version 23H2"),
    ("https://www.catalog.update.microsoft.com/Search.aspx?q=SQL+Server+2016+Service+Pack+3","SQL Server 2016","SP3 (13.0.6419.1)"),
    ("https://www.catalog.update.microsoft.com/Search.aspx?q=SQL+Server+2019","SQL Server 2019","RTM"),
    ("https://www.catalog.update.microsoft.com/Search.aspx?q=Microsoft+Office+2016+32-bit","Office 2016","x32 bits")
] 

def catalog_scrape(urls, cutoff_date, system1_csv, patchstatus_csv, header_mappings=None, remove_columns=None, filter_terms=None, screenshot=True,progress=(0,"")):
    progress.put((0,"Initializing Selenium..."))
    cutoff_date = datetime.strptime(cutoff_date, '%Y-%m-%d')
    service = Service(executable_path="chromedriver.exe")
    driver = webdriver.Chrome(service=service)

    sys1AllRows = []

    try:
        for index, (url, label, version) in enumerate(urls,start=1):
            
            progress.put((index * 2,f"Getting {label} patches..."))

            driver.get(url)
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "ctl00_catalogBody_updateMatches")))
            driver.find_element(By.ID, "ctl00_catalogBody_updateMatches_ctl02_dateHeaderLink").click()

            if screenshot:
                driver.save_screenshot(f'Screenshots/{label}_{datetime.now().strftime("%Y_%m_%d")}.png')
                progress.put((index * 2 + 2,f"Taking {label} screenshot..."))
                time.sleep(0.5)

            table = driver.find_element(By.ID, "ctl00_catalogBody_updateMatches")
            rows = table.find_elements(By.TAG_NAME, "tr")

            header = [th.text.strip() for th in rows[0].find_elements(By.TAG_NAME, 'th')]
            header[0] = '#'

            column_indices = {name: index for index, name in enumerate(header)}

            if header_mappings:
                for key, value in header_mappings.items():
                    if key in column_indices and value in column_indices:
                        header[column_indices[key]], header[column_indices[value]] = header[column_indices[value]], header[column_indices[key]]

            if remove_columns:
                header = [col for col in header if col not in remove_columns]

            system1_rows = []
            patchstatus_rows = []
            patchNumber = 1

            for row in rows[1:]:
                cells = row.find_elements(By.TAG_NAME, 'td')
                if len(cells) > 0:
                    title_text = cells[column_indices['Title']].text
                    date_str = cells[column_indices['Last Updated']].text

                    try:
                        row_date = datetime.strptime(date_str, '%m/%d/%Y')
                        if row_date >= cutoff_date:
                            kb_search = re.search(r"KB\d{6,}", title_text)
                            kb_number = kb_search.group() if kb_search else "No KB Number"

                            new_row = [patchNumber] + [cells[column_indices[col]].text for col in header[1:] if col in column_indices] + \
                                      ["Microsoft", label, version, f"{kb_number}", f"https://support.microsoft.com/kb/{kb_number}" if kb_search else "No Link"]
                            patchNumber += 1

                            if filter_terms and any(term.lower() in title_text.lower() for term in filter_terms):
                                patchstatus_rows.append(new_row)
                                patchNumber -= 1
                            else:
                                system1_rows.append(new_row)
                                sys1AllRows.append(new_row)
                    except ValueError:
                        print(f"Invalid date format found in row {title_text} {date_str}")

            header.extend(['Vendor Name', 'Vendor Product', 'Model/Version', 'Patch Name', 'Patch Link'])
            header[1] = 'Patch Description'
            header[2] = 'Release Date'
            header[3] = 'Update Type'

    finally:
        progress.put((30 ,f"Building catalog DataFrame..."))
        sys1df = pd.DataFrame(data=sys1AllRows, columns=header)
        progress.put((40 ,f"Closing Selenium..."))
        driver.quit()
        return sys1df

def map_applicability(row):
    if row.iloc[4] == "KB890830":
        msrtVer = row.loc["Patch Description"]
        msrtDecimal = int(msrtVer[46:52][3:])
        row.iloc[[13,14,15,16,17,18,19,20,21,22,23,24,25,26,27,28,29,30]] = ''
        row.loc[["#"]] = "1"
        row.loc[["Vendor Product"]] = "Windows (All)"
        row.loc[["Model/Version"]] = f"{msrtVer[46:52]}"
        row.loc[["Comment"]] = f"Replaces KB890830 {msrtVer[46:49]}{msrtDecimal-1} ({(datetime.today().replace(day=1) - timedelta(days=1)).strftime("%b-%y")})"
    elif row.iloc[2] == "Windows Server 2016":
        row.iloc[[15,16,20,21,22,23,24,25,26,27,28,29,30]] = 'N/A'
    elif row.iloc[2] == "Windows Server 2019":
        row.iloc[[13,14,17,18,19,25,26,27,28,29,30]] = 'N/A'
    elif row.iloc[2] == "Windows Server 2016 (2024)":
        row.iloc[[15,16,20,21,22,23,24,25,26,27,28,29,30]] = 'N/A'
    elif row.iloc[2] == "Windows Server 2019 (2024)":
        row.iloc[[13,14,17,18,19,25,26,27,28,29,30]] = 'N/A'
    elif row.iloc[2] == "Windows Server 2022":
        row.iloc[[13,14,15,16,17,18,19,20,21,22,23,24,27,28,29,30]] = 'N/A'
    elif row.iloc[2] == "Windows 10":
        row.iloc[[13,14,15,16,17,18,19,20,21,22,23,24,25,26,29,30]] = 'N/A'
        row.iloc[[11]] = 'Also applies for Windows 10 Version 21H2 for x64.'
    elif row.iloc[2] == "Windows 11":
        row.iloc[[13,14,15,16,17,18,19,20,21,22,23,24,25,26,27,28]] = 'N/A'
        row.iloc[[11]] = 'Also applies for Windows 11 Version 22H2 for x64.'
    elif row.iloc[2] == "SQL Server 2016":
        row.iloc[[13,14,17,18,19,20,21,22,23,24,25,26,27,29,30]] = 'N/A'
    elif row.iloc[2] == "SQL Server 2019":
        row.iloc[[15,16,17,18,19,20,21,22,23,24,25,26,28,29,30]] = 'N/A'
    elif row.iloc[2] == "Office 2016":
        row.iloc[[13,14,19,25,26,27,28,29,30]] = 'N/A'
        
    return row

def build_excel(ppw,takescreen,progress_queue):

    progress_queue.put((0,"Gathering patches from Catalog..."))

    partialCatalogDf = catalog_scrape(
        urls=catalog_urls,
        cutoff_date=ppw,
        system1_csv=f"system1_{datetime.now().strftime('%Y_%m_%d')}.csv",
        patchstatus_csv=f"patch_status_{datetime.now().strftime('%Y_%m_%d')}.csv",
        header_mappings={'Classification': 'Last Updated'},
        remove_columns=['Products', 'Version', 'Size', 'Download'],
        filter_terms=[
            'Preview', 'Dynamic', 'Azure Stack HCI', 'Office 2019', 'Access', 'Project',
            'Outlook', 'PowerPoint', 'Visio', 'Publisher', '3.5, 4.8 and 4.8.1', '3.5, 4.7.2 and 4.8',
            'SQL Server 2016 Service Pack 3 CU', 'SQL Server 2019 RTM GDR',
            '.NET Framework 3.5 and 4.8.1 for Windows 10 Version 22H2 for x64'
        ],
        screenshot=takescreen, # Set to true if screenshots are needed
        progress=progress_queue
    )

    # Catalog DataFrame without duplicates
    catalogDf = partialCatalogDf.drop_duplicates(subset=["Patch Name"], keep='last')

    progress_queue.put((45,"Building Headers..."))
    time.sleep(1)
    # Catalog DataFrame headers
    header_list = ["#", "Vendor Name", "Vendor Product", "Model/Version", "Patch Name", "Patch Description", "Patch Link", "Release Date", "Update Type", "Approved by BN", "Test Status", "Comment"]
    catalogDf = catalogDf.reindex(columns=header_list)

    progress_queue.put((50,"Building Applicability Table..."))
    time.sleep(1)

    # Systems applicability table headers
    systems_headers = ["","Server 2016 Classic 6.98 McAfee","Server 2016 Classic 6.98 Symantec","Server 2019 Classic 6.97 McAfee","Server 2019 Classic 6.97 Symantec"
                       ,"Server 2016 Evo 24.1 McAfee","Server 2016 Evo 24.1 Symantec","Server 2016 Evo 24.1 Windows Defender"
                       ,"Server 2019 Evo 23.1 McAfee","Server 2019 Evo 23.1 Symantec"
                       ,"Server 2019 Evo 23.2 McAfee","Server 2019 Evo 23.2 Symantec","Server 2019 Evo 23.2 Windows Defender"
                       ,"Server 2022 Evo 22.2 McAfee","Server 2022 Evo 22.2 Symantec","Wind10 21H2 Classic 6.98 Evo versions: 22.X - 23.1"
                       ,"Win10 22H2 Classic 6.97 Evo 23.2","Win11 22H2 Evo 23.1","Win11 23H2 Evo 24.1"]
    systemsDf = pd.DataFrame(columns=systems_headers) # Build System applicability DataFarame
    concatDf = pd.concat([catalogDf,systemsDf], axis=1) # Concatenate with Catalog Dataframe

    # Column numbers (0 indexed):
    # M - N - O - P - Q - R - S - T - U - V - W - X - Y - Z - AA - AB - AC - AD - AE
    # 12- 13- 14- 15- 16- 17- 18- 19- 20- 21- 22- 23- 24- 25- 26- 27 - 28 - 29 - 30


    finalDf = concatDf.apply(map_applicability, axis=1)

    progress_queue.put((75,"Writing Excel File..."))
    time.sleep(1)

    # Save DataFrame to Excel using openpyxl engine
    with pd.ExcelWriter("Spreadsheets/System 1 Patches.xlsx", engine="openpyxl") as writer:
        finalDf.to_excel(writer, sheet_name="September-2024", index=False)

        workbook = writer.book
        worksheet = workbook["September-2024"]

        # Set specific column and row dimensions
        worksheet.column_dimensions["A"].width = 3
        worksheet.column_dimensions["B"].width = 7
        worksheet.column_dimensions["C"].width = 11
        worksheet.column_dimensions["D"].width = 12
        worksheet.column_dimensions["E"].width = 11
        worksheet.column_dimensions["F"].width = 45
        worksheet.column_dimensions["G"].width = 30
        worksheet.column_dimensions["H"].width = 10
        worksheet.column_dimensions["I"].width = 11
        worksheet.column_dimensions["J"].width = 7
        worksheet.column_dimensions["K"].width = 7
        worksheet.column_dimensions["L"].width = 35
        worksheet.column_dimensions["M"].width = 1
        worksheet.row_dimensions[1].height = 46.5
        worksheet.row_dimensions[2].height = 46.5
        worksheet.row_dimensions[3].height = 34.5

        # Define the styles
        header_fill = PatternFill(fill_type="solid", start_color="DDEBF7", end_color="DDEBF7")
        header_font = Font(name='Calibri', color="000000", size=8, bold=True)
        center_alignment = Alignment(horizontal="center", vertical="center",wrap_text=True)

        calibri_font_8 = Font(name='Calibri', size=8)
        calibri_hyperlink = Font(name='Calibri', size=8, color="0000FF", underline="single")
        green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        yellow_fill = PatternFill(start_color="ffff00", end_color="ffff00", fill_type="solid")
        darkred_fill = PatternFill(start_color="c00000", end_color="c00000", fill_type="solid")
        darkgreen_fill = PatternFill(start_color="00b050", end_color="00b050", fill_type="solid")
        darkblue_fill = PatternFill(start_color="8ea9db", end_color="8ea9db", fill_type="solid")

        # Define thin border for all cells
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"), 
            top=Side(style="thin"), bottom=Side(style="thin")
        )

        progress_queue.put((95,"Styling Excel File..."))
        time.sleep(1)

        # Style header row
        for cell in worksheet[1]:
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = thin_border
            if cell.column in [14, 15, 26, 27, 28]:
                cell.fill = yellow_fill
            elif cell.column in [16,17,23,24,25,29]:
                cell.fill = darkred_fill
            elif cell.column in [18,19,20,31]:
                cell.fill = darkgreen_fill
            elif cell.column in [21,22,30]:
                cell.fill = darkblue_fill
            else:
                cell.fill = header_fill
            
        # Style separator line
        for row in worksheet.iter_rows(min_col=13,max_col=13):
            for cell in row:
                cell.fill = black_fill

        # Style all other rows and apply hyperlink/green fill for column E
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                # Set font for all cells
                cell.font = calibri_font_8

                # Apply borders to all cells
                cell.border = thin_border

                # Check if cell contains a URL and convert to hyperlink
                if isinstance(cell.value, str) and cell.value.startswith("http"):
                    cell.hyperlink = cell.value  # Convert to hyperlink
                    cell.font = calibri_hyperlink  # Apply hyperlink font

                # Apply green fill for column E (excluding header)
                if cell.column == 5:  # Column E (Excel is 1-indexed)
                    cell.fill = green_fill

                # Set alignment: 
                # - Center alignment for all cells horizontally and vertically
                # - For columns F and G: vertically centered but horizontally left-aligned
                if cell.column in [6, 7, 12]:  # Columns F, G and L
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                else:
                    cell.alignment = center_alignment  # Center horizontally and vertically for all other columns
        progress_queue.put((100,"Finalizing..."))
        time.sleep(1)